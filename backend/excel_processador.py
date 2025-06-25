import openpyxl
from openpyxl.styles import NumberFormatDescriptor

def localizar_e_substituir_celulas(data_excel_file_path, macro_excel_file_path):
    """
    Localiza valores em uma planilha de dados e os substitui com base em um
    mapeamento definido em um arquivo de macro. Após a substituição, formata
    a célula para o formato de data 'DD/MM/YYYY'.

    Args:
        data_excel_file_path (str): O caminho para o arquivo Excel de dados (ex: 'resultado.xlsx').
        macro_excel_file_path (str): O caminho para o arquivo Excel da Macro (ex: 'Macro - Troca de Data.xlsm').

    Raises:
        FileNotFoundError: Se um dos arquivos especificados não for encontrado.
        KeyError: Se uma das abas ('Sheet1' ou 'Planilha3') não for encontrada.
        Exception: Para outros erros inesperados durante o processamento.
    """
    try:
        #Carrega o workbook de dados e acessa a aba principal
        data_workbook = openpyxl.load_workbook(data_excel_file_path)
        ws_vendas = data_workbook['Sheet1']

        #Carrega o workbook de macro para ler o processamento
        macro_workboook = openpyxl.load_workbook(macro_excel_file_path, )
        ws_planilha3 = macro_workboook['Planilha3']

        #Cria um dicionario para armazenar os pares de procura e substituição por ser mais eficiente
        mapa_substituicoes = {} 
        for i in range (2, 33): 
            valor_procurado = ws_planilha3.cell(row = i, column = 6).value  # Coluna F: Valor a ser procurado na planilha de dados.
            valor_substituto = ws_planilha3.cell(row = i, column = 7).value  # Coluna G: Novo valor que substituirá o valor encontrado.

            #verifica se os valor não sao vazios antes de adiciona-los
            if valor_procurado is not None and valor_substituto is not None:
                mapa_substituicoes[str(valor_procurado)] = valor_substituto 

        macro_workboook.close() #libera o arquivo da macro

        for row in ws_vendas.iter_rows(): #percorre todas as linhas com o iter_rows()
            for cell in row: #para cada linha, percorre cada celula
                if cell.value is None:
                    continue

                valor_cell_str = str(cell.value)

                #verifica se valor_cell não esta vazio e se seu valor, convertido para string, esta na lista de valores a serem procurados
                if valor_cell_str in mapa_substituicoes:
                    cell.value = mapa_substituicoes[valor_cell_str]
                    cell.number_format = 'DD/MM/YYYY'

        data_workbook.save(data_excel_file_path)
        print(f"Substituições de celulas concluídas e salva em: {data_excel_file_path}")

    except FileNotFoundError as e:
        print(f"Erro: Um dos arquivos Excel não foi encontrado: {e}")
        raise
    except KeyError as e:
        print(f"Erro: Uma das Abas não foi encontrada nas planilhas, verifique o excel e o macro: {e}")
        raise
    except Exception as e:
        print(f"Ocorreu um erro ao processar o excel: {e}")
        raise

    
if __name__ == "__main__":
    pass
