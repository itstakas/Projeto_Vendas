from openpyxl import load_workbook
from datetime import datetime
import os

def adicionar_clientes_manualmente(caminho_arquivo: str):   

    if not os.path.exists(caminho_arquivo):
        print(f"Arquivo não encontrado")
        return

    wb = load_workbook(caminho_arquivo)
    ws = wb.active

    novos_clientes = [
        {"NOME": "NAIANNNY GUERRIERI", "VENDEDOR_TELE": "Lucas Dos Santos Delgado", "DATA_CONTRATO": "06/06/2025"},#datetime.now().strftime('%d/%m/%Y')
        {"NOME": "LUIZ HENRIQUE AGUIAR", "VENDEDOR_TELE": "Sheila Licia Nascimento Silva", "DATA_CONTRATO": "06/06/2025"}#datetime.now().strftime('%d/%m/%Y')
    ]

    headers = {cell.value: idx for idx, cell in enumerate(ws[1])}
    nome_col = headers.get("NOME")
    vendedor_col = headers.get("VENDEDOR_TELE")
    data_col = headers.get("DATA_CONTRATO")

    if nome_col is None or vendedor_col is None or data_col is None:
        print("Colunas não encontradas")
        return
    
    for cliente in novos_clientes:
        nova_linha = [""] * len(headers)
        nova_linha[nome_col] = cliente["NOME"]
        nova_linha[vendedor_col] = cliente["VENDEDOR_TELE"]
        nova_linha[data_col] = cliente["DATA_CONTRATO"]
        ws.append(nova_linha)

    wb.save(caminho_arquivo)
    print(f"Clientes adicionados com sucesso em {caminho_arquivo}")
